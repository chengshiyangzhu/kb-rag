"""XLSX parser using openpyxl.

Each worksheet becomes its own :class:`Document`. The sheet content is
serialised as a Markdown table so downstream structural chunkers can treat it as
an atomic block.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

from app.ingest.parsers.base import Parser, _make_doc_id
from app.models.document import Document, Metadata


def _rows_to_markdown(rows: list[tuple[Any, ...]]) -> str:
    """Serialise a sequence of worksheet rows into a Markdown table.

    Args:
        rows: A list of row tuples as returned by ``ws.iter_rows(values_only=True)``.

    Returns:
        A Markdown table string. Empty input yields an empty string.
    """
    if not rows:
        return ""
    ncols = max(len(row) for row in rows)
    if ncols == 0:
        return ""
    lines: list[str] = []
    header = list(rows[0]) + [None] * (ncols - len(rows[0]))
    lines.append("| " + " | ".join(_cell_str(c) for c in header) + " |")
    lines.append("| " + " | ".join("---" for _ in range(ncols)) + " |")
    for row in rows[1:]:
        cells = list(row) + [None] * (ncols - len(row))
        lines.append("| " + " | ".join(_cell_str(c) for c in cells[:ncols]) + " |")
    return "\n".join(lines)


def _cell_str(value: Any) -> str:
    """Render a cell value as a concise string.

    Args:
        value: The raw cell value (may be ``None``).

    Returns:
        The string representation of the value, or an empty string for ``None``.
    """
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


class XlsxParser(Parser):
    """Parse an ``.xlsx``/``.xls`` file into one :class:`Document` per sheet."""

    def parse(self, file_path: Path) -> list[Document]:
        """Extract each worksheet as a Markdown-table Document.

        Args:
            file_path: Path to the spreadsheet file.

        Returns:
            A list of :class:`Document` instances, one per worksheet.
        """
        from openpyxl import load_workbook

        wb = load_workbook(str(file_path), data_only=True, read_only=True)
        doc_id = _make_doc_id(file_path, None)
        docs: list[Document] = []
        try:
            for sheet_idx, sheet_name in enumerate(wb.sheetnames):
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                text = _rows_to_markdown(rows)
                metadata = Metadata(
                    source=file_path.name,
                    page=None,
                    sheet=sheet_name,
                    doc_id=doc_id,
                )
                docs.append(
                    Document(
                        id=_make_doc_id(file_path, sheet_idx),
                        text=text,
                        metadata=metadata,
                    )
                )
        finally:
            wb.close()
        return docs
